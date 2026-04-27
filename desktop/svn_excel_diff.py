#!/usr/bin/env python3
"""SVN Excel Diff Tool

Compares the working copy of Excel files against their SVN base version,
producing cell-level diffs across all sheets.

Rows are matched by the primary key (first column) so that inserted/deleted
rows are correctly identified instead of causing cascading false diffs.

Usage:
    python svn_excel_diff.py <svn_working_copy_path> [--file <specific_file>] [--html]

Output: Structured text report, or an HTML visual diff opened in browser.
"""

import argparse
import difflib
import html as html_mod
import json
import os
import shutil
import string
import subprocess
import sys
import tempfile
import webbrowser
from collections import OrderedDict
from datetime import datetime

import openpyxl
import xlrd


SVN_CANDIDATES = (
    "/opt/homebrew/bin/svn",
    "/usr/local/bin/svn",
    "/usr/bin/svn",
    "/opt/local/bin/svn",
)


def _run_subprocess(*args, **kwargs):
    """Run subprocesses without flashing console windows in packaged Windows GUI."""
    if os.name == "nt":
        kwargs.setdefault("creationflags", subprocess.CREATE_NO_WINDOW)
    return subprocess.run(*args, **kwargs)


def get_svn_command():
    """Return an svn executable path that works from shells and macOS .app launches."""
    svn = shutil.which("svn")
    if svn:
        return svn
    for candidate in SVN_CANDIDATES:
        if os.path.isfile(candidate) and os.access(candidate, os.X_OK):
            return candidate
    return "svn"


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------

def read_excel_to_rows(filepath):
    """Read an Excel file into a row-oriented structure.

    Returns:
        {sheet_name: {
            "headers": [col0_name, col1_name, ...],
            "rows": OrderedDict{key_value: [col0, col1, ...], ...}
        }}
    """
    ext = os.path.splitext(filepath)[1].lower()
    result = {}

    if ext == ".xls":
        wb = xlrd.open_workbook(filepath)
        for name in wb.sheet_names():
            sh = wb.sheet_by_name(name)
            if sh.nrows == 0:
                result[name] = {"headers": [], "rows": OrderedDict()}
                continue
            headers = []
            for c in range(sh.ncols):
                val = sh.cell_value(0, c)
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                headers.append(str(val) if val != "" else f"col_{c}")
            rows = OrderedDict()
            for r in range(1, sh.nrows):
                row_data = []
                for c in range(sh.ncols):
                    val = sh.cell_value(r, c)
                    cell_type = sh.cell_type(r, c)
                    if cell_type == xlrd.XL_CELL_EMPTY:
                        val = ""
                    elif isinstance(val, float) and val == int(val):
                        val = int(val)
                    row_data.append(val)
                key = row_data[0] if row_data else r
                if key in rows:
                    key = f"{key}_row{r + 1}"
                rows[key] = row_data
            result[name] = {"headers": headers, "rows": rows}

    elif ext == ".xlsx":
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            all_rows = list(ws.iter_rows())
            if not all_rows:
                result[name] = {"headers": [], "rows": OrderedDict()}
                continue
            max_col = max(len(row) for row in all_rows)
            headers = []
            for c in range(max_col):
                val = all_rows[0][c].value if c < len(all_rows[0]) else None
                if val is None:
                    val = ""
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                headers.append(str(val) if val != "" else f"col_{c}")
            rows = OrderedDict()
            for r_idx, row in enumerate(all_rows[1:], start=2):
                row_data = []
                for c in range(max_col):
                    val = row[c].value if c < len(row) else None
                    if val is None:
                        val = ""
                    elif isinstance(val, float) and val == int(val):
                        val = int(val)
                    row_data.append(val)
                key = row_data[0] if row_data else r_idx
                if key in rows:
                    key = f"{key}_row{r_idx}"
                rows[key] = row_data
            result[name] = {"headers": headers, "rows": rows}
        wb.close()
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    return result


# ---------------------------------------------------------------------------
# SVN helpers
# ---------------------------------------------------------------------------

def get_svn_base(filepath):
    try:
        result = _run_subprocess(
            [get_svn_command(), "cat", filepath], capture_output=True, check=True,
        )
        ext = os.path.splitext(filepath)[1]
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
            f.write(result.stdout)
            return f.name
    except subprocess.CalledProcessError:
        return None


def get_svn_status(directory):
    result = _run_subprocess(
        [get_svn_command(), "status", directory], capture_output=True, text=True,
    )
    files = []
    for line in result.stdout.strip().split("\n"):
        if not line.strip():
            continue
        status = line[0]
        filepath = line[8:].strip() if len(line) > 8 else line[1:].strip()
        files.append((status, filepath))
    return files


def get_svn_info(filepath):
    """Get SVN revision info for display."""
    result = _run_subprocess(
        [get_svn_command(), "info", filepath], capture_output=True, text=True,
    )
    info = {}
    for line in result.stdout.strip().split("\n"):
        if ":" in line:
            k, v = line.split(":", 1)
            info[k.strip()] = v.strip()
    return info


# ---------------------------------------------------------------------------
# Diff logic  — produces a unified row list for side-by-side rendering
# ---------------------------------------------------------------------------

def _row_content_key(row, skip_col=0):
    """Create a hashable content signature from a row, skipping the ID column.

    This is used by SequenceMatcher to align rows by content similarity
    rather than by the (unstable) primary key / ID column.
    """
    return tuple(row[c] for c in range(len(row)) if c != skip_col)


def _compare_rows(base_row, work_row):
    """Compare two aligned rows cell-by-cell, return set of changed col indices."""
    max_c = max(len(base_row), len(work_row))
    changed = set()
    for c in range(max_c):
        ov = base_row[c] if c < len(base_row) else ""
        nv = work_row[c] if c < len(work_row) else ""
        if ov != nv:
            changed.add(c)
    return changed


def build_unified_diff(base_data, work_data):
    """Build a unified row list suitable for side-by-side HTML rendering.

    Uses difflib.SequenceMatcher on row content (excluding the first column)
    so that rows with renumbered IDs are correctly matched as modifications
    instead of delete+add.

    Returns per-file structure:
    {
      sheet_name: {
        "headers": [...],
        "base_headers": [...],
        "rows": [
            {
              "type": "equal" | "added" | "deleted" | "modified",
              "key": ...,
              "base": [values...] | None,
              "work": [values...] | None,
              "changed_cols": set of col indices (for modified rows),
            }, ...
        ]
      }
    }
    """
    all_sheet_names = []
    seen = set()
    for name in list(base_data.keys()) + list(work_data.keys()):
        if name not in seen:
            all_sheet_names.append(name)
            seen.add(name)

    result = {}
    for sheet_name in all_sheet_names:
        base_info = base_data.get(sheet_name, {"headers": [], "rows": OrderedDict()})
        work_info = work_data.get(sheet_name, {"headers": [], "rows": OrderedDict()})
        base_rows = base_info["rows"]
        work_rows = work_info["rows"]
        headers = work_info["headers"] if work_info["headers"] else base_info["headers"]
        base_headers = base_info["headers"]

        base_list = list(base_rows.values())
        work_list = list(work_rows.values())

        # Build content signatures (skip column 0 = ID) for sequence matching
        base_sigs = [_row_content_key(r, skip_col=0) for r in base_list]
        work_sigs = [_row_content_key(r, skip_col=0) for r in work_list]

        matcher = difflib.SequenceMatcher(None, base_sigs, work_sigs, autojunk=False)
        unified = []

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == "equal":
                # Rows matched by content — check if ID or any cell actually changed
                for bi, wi in zip(range(i1, i2), range(j1, j2)):
                    base_row = base_list[bi]
                    work_row = work_list[wi]
                    changed = _compare_rows(base_row, work_row)
                    rtype = "modified" if changed else "equal"
                    unified.append({
                        "type": rtype,
                        "key": work_row[0] if work_row else bi,
                        "base": base_row, "work": work_row,
                        "changed_cols": changed,
                    })

            elif tag == "delete":
                for bi in range(i1, i2):
                    base_row = base_list[bi]
                    unified.append({
                        "type": "deleted",
                        "key": base_row[0] if base_row else bi,
                        "base": base_row, "work": None,
                        "changed_cols": set(),
                    })

            elif tag == "insert":
                for wi in range(j1, j2):
                    work_row = work_list[wi]
                    unified.append({
                        "type": "added",
                        "key": work_row[0] if work_row else wi,
                        "base": None, "work": work_row,
                        "changed_cols": set(),
                    })

            elif tag == "replace":
                # Rows in this block differ. Try to pair them by similarity.
                base_block = list(range(i1, i2))
                work_block = list(range(j1, j2))

                # Pair up rows greedily by best content similarity
                paired_base = set()
                paired_work = set()
                pairs = []

                if len(base_block) <= 50 and len(work_block) <= 50:
                    # For small blocks, find best pairs by similarity ratio
                    scores = []
                    for bi in base_block:
                        for wi in work_block:
                            # Compare content excluding ID column
                            bs = base_sigs[bi]
                            ws = work_sigs[wi]
                            # Quick similarity: count matching fields
                            match_count = sum(1 for a, b in zip(bs, ws) if a == b)
                            total = max(len(bs), len(ws))
                            ratio = match_count / total if total > 0 else 0
                            if ratio > 0.4:  # At least 40% similar to consider a match
                                scores.append((ratio, bi, wi))
                    scores.sort(reverse=True)

                    for ratio, bi, wi in scores:
                        if bi in paired_base or wi in paired_work:
                            continue
                        paired_base.add(bi)
                        paired_work.add(wi)
                        pairs.append((bi, wi))

                # Emit in order: process base_block and work_block positions
                # interleaving paired modifications, unpaired deletes, and unpaired adds
                pair_map_b = {bi: wi for bi, wi in pairs}
                pair_map_w = {wi: bi for bi, wi in pairs}

                # Walk through both blocks in order
                bi_idx = 0
                wi_idx = 0
                while bi_idx < len(base_block) or wi_idx < len(work_block):
                    bi = base_block[bi_idx] if bi_idx < len(base_block) else None
                    wi = work_block[wi_idx] if wi_idx < len(work_block) else None

                    if bi is not None and bi in pair_map_b:
                        # This base row is paired — emit any unpaired work rows before its pair
                        target_wi = pair_map_b[bi]
                        while wi_idx < len(work_block) and work_block[wi_idx] < target_wi:
                            wwi = work_block[wi_idx]
                            if wwi not in paired_work:
                                work_row = work_list[wwi]
                                unified.append({
                                    "type": "added",
                                    "key": work_row[0] if work_row else wwi,
                                    "base": None, "work": work_row,
                                    "changed_cols": set(),
                                })
                            wi_idx += 1

                        # Emit the paired modification
                        base_row = base_list[bi]
                        work_row = work_list[target_wi]
                        changed = _compare_rows(base_row, work_row)
                        rtype = "modified" if changed else "equal"
                        unified.append({
                            "type": rtype,
                            "key": work_row[0] if work_row else bi,
                            "base": base_row, "work": work_row,
                            "changed_cols": changed,
                        })
                        bi_idx += 1
                        # Advance wi_idx past the paired work row
                        if wi_idx < len(work_block) and work_block[wi_idx] == target_wi:
                            wi_idx += 1

                    elif bi is not None and bi not in paired_base:
                        # Unpaired base row — deleted
                        base_row = base_list[bi]
                        unified.append({
                            "type": "deleted",
                            "key": base_row[0] if base_row else bi,
                            "base": base_row, "work": None,
                            "changed_cols": set(),
                        })
                        bi_idx += 1

                    elif wi is not None and wi not in paired_work:
                        # Unpaired work row — added
                        work_row = work_list[wi]
                        unified.append({
                            "type": "added",
                            "key": work_row[0] if work_row else wi,
                            "base": None, "work": work_row,
                            "changed_cols": set(),
                        })
                        wi_idx += 1

                    else:
                        # Skip already handled
                        if bi is not None:
                            bi_idx += 1
                        if wi is not None:
                            wi_idx += 1

                # Emit remaining unpaired work rows
                while wi_idx < len(work_block):
                    wwi = work_block[wi_idx]
                    if wwi not in paired_work:
                        work_row = work_list[wwi]
                        unified.append({
                            "type": "added",
                            "key": work_row[0] if work_row else wwi,
                            "base": None, "work": work_row,
                            "changed_cols": set(),
                        })
                    wi_idx += 1

        result[sheet_name] = {
            "headers": headers,
            "base_headers": base_headers,
            "rows": unified,
        }

    return result


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SVN Diff — $title</title>
<style>
:root {
  --bg: #1e1e1e;
  --bg-surface: #252526;
  --bg-header: #2d2d2d;
  --text: #d4d4d4;
  --text-dim: #808080;
  --border: #404040;
  --added-bg: #1e3a1e;
  --added-line: #2ea04333;
  --added-text: #3fb950;
  --deleted-bg: #3a1e1e;
  --deleted-line: #f8514933;
  --deleted-text: #f85149;
  --modified-bg: #3a3a1e;
  --modified-cell: #d29922;
  --modified-cell-bg: rgba(210,153,34,0.15);
  --equal-bg: transparent;
  --tab-active: #1e1e1e;
  --tab-inactive: #2d2d2d;
  --tab-hover: #383838;
  --accent: #569cd6;
  --scrollbar-thumb: #555;
}
* { margin:0; padding:0; box-sizing:border-box; }
body {
  font-family: 'Menlo', 'Consolas', 'Courier New', monospace;
  font-size: 12px;
  background: var(--bg);
  color: var(--text);
  line-height: 1.5;
}
.toolbar {
  position: sticky; top:0; z-index:100;
  background: var(--bg-header);
  border-bottom: 1px solid var(--border);
  padding: 8px 16px;
  display: flex;
  align-items: center;
  gap: 16px;
}
.toolbar h1 { font-size:14px; font-weight:600; color:var(--accent); }
.toolbar .meta { color:var(--text-dim); font-size:11px; }
.stats {
  display:flex; gap:12px; margin-left:auto;
}
.stat {
  padding: 2px 8px;
  border-radius: 3px;
  font-size: 11px;
  font-weight: 600;
}
.stat.added { background:var(--added-bg); color:var(--added-text); }
.stat.deleted { background:var(--deleted-bg); color:var(--deleted-text); }
.stat.modified { background:var(--modified-bg); color:var(--modified-cell); }

.file-section {
  margin: 12px;
  border: 1px solid var(--border);
  border-radius: 6px;
  overflow: hidden;
}
.file-header {
  background: var(--bg-header);
  padding: 8px 16px;
  border-bottom: 1px solid var(--border);
  display: flex;
  align-items: center;
  gap: 8px;
}
.file-header .filename {
  font-weight: 600;
  color: var(--text);
}
.file-header .badge {
  font-size: 10px;
  padding: 1px 6px;
  border-radius: 3px;
  font-weight: 600;
}
.badge.M { background:var(--modified-bg); color:var(--modified-cell); }
.badge.A { background:var(--added-bg); color:var(--added-text); }
.badge.D { background:var(--deleted-bg); color:var(--deleted-text); }

/* Sheet tabs */
.sheet-tabs {
  display: flex;
  background: var(--bg-header);
  border-bottom: 1px solid var(--border);
  overflow-x: auto;
}
.sheet-tab {
  padding: 6px 16px;
  cursor: pointer;
  border-bottom: 2px solid transparent;
  color: var(--text-dim);
  white-space: nowrap;
  font-size: 12px;
  user-select: none;
  transition: all 0.15s;
}
.sheet-tab:hover { background: var(--tab-hover); color: var(--text); }
.sheet-tab.active {
  background: var(--tab-active);
  color: var(--accent);
  border-bottom-color: var(--accent);
}
.sheet-tab .tab-badge {
  font-size: 10px;
  margin-left: 4px;
  padding: 0 4px;
  border-radius: 3px;
}

.sheet-content { display:none; }
.sheet-content.active { display:block; }

/* Filter buttons */
.filter-bar {
  background: var(--bg-surface);
  padding: 6px 16px;
  border-bottom: 1px solid var(--border);
  display: flex;
  gap: 8px;
  align-items: center;
}
.filter-btn {
  background: var(--bg-header);
  border: 1px solid var(--border);
  color: var(--text-dim);
  padding: 2px 10px;
  border-radius: 3px;
  cursor: pointer;
  font-size: 11px;
  font-family: inherit;
  transition: all 0.15s;
}
.filter-btn:hover { border-color: var(--accent); color: var(--text); }
.filter-btn.active { border-color: var(--accent); color: var(--accent); background: rgba(86,156,214,0.1); }
.filter-label { color: var(--text-dim); font-size: 11px; }

/* Diff table */
.diff-wrapper {
  overflow-x: auto;
}
.diff-table {
  width: 100%;
  border-collapse: collapse;
  table-layout: auto;
}
.diff-table th {
  position: sticky;
  top: 0;
  background: var(--bg-header);
  color: var(--text-dim);
  font-weight: 600;
  font-size: 11px;
  text-align: left;
  padding: 6px 8px;
  border-bottom: 2px solid var(--border);
  border-right: 1px solid var(--border);
  white-space: nowrap;
  z-index: 10;
}
.diff-table th.side-label {
  text-align: center;
  font-size: 12px;
  color: var(--accent);
  letter-spacing: 1px;
}
.diff-table th.separator {
  width: 3px;
  min-width: 3px;
  padding: 0;
  background: var(--accent);
  border: none;
}
.diff-table td {
  padding: 3px 8px;
  border-right: 1px solid var(--border);
  border-bottom: 1px solid rgba(255,255,255,0.04);
  white-space: nowrap;
  max-width: 300px;
  overflow: hidden;
  text-overflow: ellipsis;
  vertical-align: top;
}
.diff-table td.separator {
  width: 3px;
  min-width: 3px;
  padding: 0;
  background: var(--accent);
  border: none;
}
.diff-table td.line-num {
  color: var(--text-dim);
  text-align: right;
  user-select: none;
  width: 36px;
  min-width: 36px;
  padding-right: 6px;
  font-size: 11px;
}

/* Row types */
tr.row-equal td { background: var(--equal-bg); }
tr.row-added td { background: var(--added-bg); }
tr.row-added td.left-side { background: var(--bg-surface); color: var(--text-dim); }
tr.row-deleted td { background: var(--deleted-bg); }
tr.row-deleted td.right-side { background: var(--bg-surface); color: var(--text-dim); }
tr.row-modified td { background: var(--equal-bg); }
tr.row-modified td.cell-changed { background: var(--modified-cell-bg); }

/* Gutter indicators */
td.gutter {
  width: 4px;
  min-width: 4px;
  padding: 0;
  border-right: none;
}
tr.row-added td.gutter.right { background: var(--added-text); }
tr.row-deleted td.gutter.left { background: var(--deleted-text); }
tr.row-modified td.gutter { background: var(--modified-cell); }

/* Cell highlight tags */
.old-val {
  background: rgba(248,81,73,0.25);
  color: var(--deleted-text);
  padding: 0 2px;
  border-radius: 2px;
}
.new-val {
  background: rgba(46,160,67,0.25);
  color: var(--added-text);
  padding: 0 2px;
  border-radius: 2px;
}

/* Scrollbar */
::-webkit-scrollbar { width:8px; height:8px; }
::-webkit-scrollbar-track { background:var(--bg); }
::-webkit-scrollbar-thumb { background:var(--scrollbar-thumb); border-radius:4px; }

/* No changes */
.no-changes {
  padding: 40px;
  text-align: center;
  color: var(--text-dim);
  font-size: 14px;
}
</style>
</head>
<body>

<div class="toolbar">
  <h1>SVN Diff</h1>
  <span class="meta">$directory &mdash; $timestamp</span>
  <div class="stats">
    <span class="stat added">+$total_added added</span>
    <span class="stat deleted">&minus;$total_deleted deleted</span>
    <span class="stat modified">~$total_modified modified</span>
  </div>
</div>

$file_sections

<script>
// Sheet tab switching
document.querySelectorAll('.sheet-tab').forEach(tab => {
  tab.addEventListener('click', () => {
    const fileId = tab.dataset.file;
    const sheetId = tab.dataset.sheet;
    document.querySelectorAll(`.sheet-tab[data-file="${fileId}"]`).forEach(t => t.classList.remove('active'));
    document.querySelectorAll(`.sheet-content[data-file="${fileId}"]`).forEach(c => c.classList.remove('active'));
    tab.classList.add('active');
    document.querySelector(`.sheet-content[data-file="${fileId}"][data-sheet="${sheetId}"]`).classList.add('active');
  });
});

// Filter buttons
document.querySelectorAll('.filter-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    btn.classList.toggle('active');
    applyFilters(btn.closest('.file-section'));
  });
});

function applyFilters(section) {
  const buttons = section.querySelectorAll('.filter-btn');
  const active = {};
  buttons.forEach(b => { active[b.dataset.type] = b.classList.contains('active'); });
  // If none active, show all
  const anyActive = Object.values(active).some(v => v);

  section.querySelectorAll('.diff-table tbody tr').forEach(row => {
    if (!anyActive) {
      row.style.display = '';
      return;
    }
    const type = row.dataset.type;
    row.style.display = active[type] ? '' : 'none';
  });
}
</script>

</body>
</html>"""


def _esc(val):
    return html_mod.escape(str(val)) if val != "" else '<span style="opacity:0.3">-</span>'


def generate_html(diff_results, directory):
    """Generate a self-contained HTML diff page.

    diff_results: list of (filepath, status, unified_diff_or_None, text_diff_or_None)
    """
    file_sections = []
    total_added = 0
    total_deleted = 0
    total_modified = 0

    for file_idx, (filepath, status, udiff, text_diff) in enumerate(diff_results):
        fname = os.path.basename(filepath)
        fid = f"f{file_idx}"

        if status == "A":
            file_sections.append(f'''
            <div class="file-section">
              <div class="file-header">
                <span class="badge A">A</span>
                <span class="filename">{_esc(fname)}</span>
              </div>
              <div class="no-changes">New file added</div>
            </div>''')
            continue
        if status == "D":
            file_sections.append(f'''
            <div class="file-section">
              <div class="file-header">
                <span class="badge D">D</span>
                <span class="filename">{_esc(fname)}</span>
              </div>
              <div class="no-changes">File deleted</div>
            </div>''')
            continue

        if text_diff:
            escaped_diff = _esc(text_diff)
            file_sections.append(f'''
            <div class="file-section">
              <div class="file-header">
                <span class="badge M">M</span>
                <span class="filename">{_esc(fname)}</span>
              </div>
              <pre style="padding:12px;overflow-x:auto;color:var(--text);">{escaped_diff}</pre>
            </div>''')
            continue

        if not udiff:
            continue

        # Build sheet tabs and content
        sheet_names = list(udiff.keys())
        tabs_html = []
        contents_html = []

        for s_idx, sheet_name in enumerate(sheet_names):
            sdata = udiff[sheet_name]
            sid = f"s{s_idx}"
            rows = sdata["rows"]
            headers = sdata["headers"]
            base_headers = sdata["base_headers"]
            num_cols = max(len(headers), len(base_headers)) if headers or base_headers else 0

            s_added = sum(1 for r in rows if r["type"] == "added")
            s_deleted = sum(1 for r in rows if r["type"] == "deleted")
            s_modified = sum(1 for r in rows if r["type"] == "modified")
            total_added += s_added
            total_deleted += s_deleted
            total_modified += s_modified

            has_changes = s_added + s_deleted + s_modified > 0
            active = "active" if s_idx == 0 else ""

            badge_parts = []
            if s_added: badge_parts.append(f'+{s_added}')
            if s_deleted: badge_parts.append(f'-{s_deleted}')
            if s_modified: badge_parts.append(f'~{s_modified}')
            badge_text = " ".join(badge_parts)
            badge_style = ""
            if not has_changes:
                badge_text = "no changes"
                badge_style = "background:var(--bg-surface);color:var(--text-dim);"
            else:
                badge_style = "background:var(--modified-bg);color:var(--modified-cell);"

            tabs_html.append(
                f'<div class="sheet-tab {active}" data-file="{fid}" data-sheet="{sid}">'
                f'{_esc(sheet_name)}'
                f'<span class="tab-badge" style="{badge_style}">{badge_text}</span>'
                f'</div>'
            )

            if not has_changes:
                contents_html.append(
                    f'<div class="sheet-content {active}" data-file="{fid}" data-sheet="{sid}">'
                    f'<div class="no-changes">No changes in this sheet</div></div>'
                )
                continue

            # Build table
            # Headers: line# | gutter | base cols... | separator | line# | gutter | work cols...
            th_base = "".join(f"<th>{_esc(base_headers[c] if c < len(base_headers) else '')}</th>" for c in range(num_cols))
            th_work = "".join(f"<th>{_esc(headers[c] if c < len(headers) else '')}</th>" for c in range(num_cols))
            thead = (
                f'<tr>'
                f'<th class="side-label" colspan="{num_cols + 2}">BASE (SVN)</th>'
                f'<th class="separator"></th>'
                f'<th class="side-label" colspan="{num_cols + 2}">WORKING COPY</th>'
                f'</tr>'
                f'<tr>'
                f'<th>#</th><th></th>{th_base}'
                f'<th class="separator"></th>'
                f'<th>#</th><th></th>{th_work}'
                f'</tr>'
            )

            tbody_rows = []
            base_line = 0
            work_line = 0
            for row in rows:
                rtype = row["type"]
                key = row["key"]
                changed = row["changed_cols"]

                if rtype == "equal":
                    base_line += 1
                    work_line += 1
                    base_data = row["base"]
                    work_data_r = row["work"]
                    left_cells = "".join(f"<td>{_esc(base_data[c] if c < len(base_data) else '')}</td>" for c in range(num_cols))
                    right_cells = "".join(f"<td>{_esc(work_data_r[c] if c < len(work_data_r) else '')}</td>" for c in range(num_cols))
                    tbody_rows.append(
                        f'<tr class="row-equal" data-type="equal">'
                        f'<td class="line-num">{base_line}</td><td class="gutter left"></td>{left_cells}'
                        f'<td class="separator"></td>'
                        f'<td class="line-num">{work_line}</td><td class="gutter right"></td>{right_cells}'
                        f'</tr>'
                    )
                elif rtype == "deleted":
                    base_line += 1
                    base_data = row["base"]
                    left_cells = "".join(f"<td>{_esc(base_data[c] if c < len(base_data) else '')}</td>" for c in range(num_cols))
                    right_cells = f'<td class="right-side" colspan="{num_cols}"></td>'
                    tbody_rows.append(
                        f'<tr class="row-deleted" data-type="deleted">'
                        f'<td class="line-num">{base_line}</td><td class="gutter left"></td>{left_cells}'
                        f'<td class="separator"></td>'
                        f'<td class="line-num right-side"></td><td class="gutter right"></td>{right_cells}'
                        f'</tr>'
                    )
                elif rtype == "added":
                    work_line += 1
                    work_data_r = row["work"]
                    left_cells = f'<td class="left-side" colspan="{num_cols}"></td>'
                    right_cells = "".join(f"<td>{_esc(work_data_r[c] if c < len(work_data_r) else '')}</td>" for c in range(num_cols))
                    tbody_rows.append(
                        f'<tr class="row-added" data-type="added">'
                        f'<td class="line-num left-side"></td><td class="gutter left"></td>{left_cells}'
                        f'<td class="separator"></td>'
                        f'<td class="line-num">{work_line}</td><td class="gutter right"></td>{right_cells}'
                        f'</tr>'
                    )
                elif rtype == "modified":
                    base_line += 1
                    work_line += 1
                    base_data = row["base"]
                    work_data_r = row["work"]
                    left_cells = []
                    right_cells = []
                    for c in range(num_cols):
                        bv = base_data[c] if c < len(base_data) else ""
                        wv = work_data_r[c] if c < len(work_data_r) else ""
                        if c in changed:
                            left_cells.append(f'<td class="cell-changed"><span class="old-val">{_esc(bv)}</span></td>')
                            right_cells.append(f'<td class="cell-changed"><span class="new-val">{_esc(wv)}</span></td>')
                        else:
                            left_cells.append(f'<td>{_esc(bv)}</td>')
                            right_cells.append(f'<td>{_esc(wv)}</td>')
                    tbody_rows.append(
                        f'<tr class="row-modified" data-type="modified">'
                        f'<td class="line-num">{base_line}</td><td class="gutter left"></td>{"".join(left_cells)}'
                        f'<td class="separator"></td>'
                        f'<td class="line-num">{work_line}</td><td class="gutter right"></td>{"".join(right_cells)}'
                        f'</tr>'
                    )

            filter_bar = (
                f'<div class="filter-bar">'
                f'<span class="filter-label">Filter:</span>'
                f'<button class="filter-btn" data-type="added">+ Added ({s_added})</button>'
                f'<button class="filter-btn" data-type="deleted">- Deleted ({s_deleted})</button>'
                f'<button class="filter-btn" data-type="modified">~ Modified ({s_modified})</button>'
                f'</div>'
            )

            table_html = (
                f'{filter_bar}'
                f'<div class="diff-wrapper">'
                f'<table class="diff-table">'
                f'<thead>{thead}</thead>'
                f'<tbody>{"".join(tbody_rows)}</tbody>'
                f'</table></div>'
            )

            contents_html.append(
                f'<div class="sheet-content {active}" data-file="{fid}" data-sheet="{sid}">'
                f'{table_html}</div>'
            )

        file_sections.append(
            f'<div class="file-section">'
            f'<div class="file-header">'
            f'<span class="badge M">M</span>'
            f'<span class="filename">{_esc(fname)}</span>'
            f'</div>'
            f'<div class="sheet-tabs">{"".join(tabs_html)}</div>'
            f'{"".join(contents_html)}'
            f'</div>'
        )

    html_content = string.Template(HTML_TEMPLATE).safe_substitute(
        title=html_mod.escape(os.path.basename(directory)),
        directory=html_mod.escape(directory),
        timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        total_added=total_added,
        total_deleted=total_deleted,
        total_modified=total_modified,
        file_sections="\n".join(file_sections),
    )
    return html_content


# ---------------------------------------------------------------------------
# Text report (kept for non-HTML mode)
# ---------------------------------------------------------------------------

def format_report(file_path, udiff, is_new=False, is_deleted=False):
    lines = []
    rel_name = os.path.basename(file_path)

    if is_new:
        lines.append(f"### [A] {rel_name}")
        lines.append("New file added\n")
        return "\n".join(lines)
    if is_deleted:
        lines.append(f"### [D] {rel_name}")
        lines.append("File deleted\n")
        return "\n".join(lines)
    if not udiff:
        return ""

    lines.append(f"### [M] {rel_name}\n")
    for sheet_name, sdata in udiff.items():
        rows = sdata["rows"]
        headers = sdata["headers"]
        added = [r for r in rows if r["type"] == "added"]
        deleted = [r for r in rows if r["type"] == "deleted"]
        modified = [r for r in rows if r["type"] == "modified"]

        parts = []
        if added: parts.append(f"+{len(added)} rows added")
        if deleted: parts.append(f"-{len(deleted)} rows deleted")
        if modified: parts.append(f"~{len(modified)} rows modified")
        if not parts:
            continue

        lines.append(f"**Sheet `{sheet_name}`**: {', '.join(parts)}\n")

        if deleted:
            lines.append("#### Deleted Rows")
            for r in deleted:
                lines.append(f"  - **ID={r['key']}**")
                for c, val in enumerate(r["base"]):
                    if val == "": continue
                    hdr = headers[c] if c < len(headers) else f"col_{c}"
                    lines.append(f"    {hdr} = `{val}`")
            lines.append("")
        if added:
            lines.append("#### Added Rows")
            for r in added:
                lines.append(f"  - **ID={r['key']}**")
                for c, val in enumerate(r["work"]):
                    if val == "": continue
                    hdr = headers[c] if c < len(headers) else f"col_{c}"
                    lines.append(f"    {hdr} = `{val}`")
            lines.append("")
        if modified:
            lines.append("#### Modified Rows")
            for r in modified:
                lines.append(f"  - **ID={r['key']}**:")
                base_row = r["base"]
                work_row = r["work"]
                for c in sorted(r["changed_cols"]):
                    hdr = headers[c] if c < len(headers) else f"col_{c}"
                    ov = base_row[c] if c < len(base_row) else ""
                    nv = work_row[c] if c < len(work_row) else ""
                    lines.append(f"    - {hdr}: `{ov or '_(empty)_'}` -> `{nv or '_(empty)_'}`")
            lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="SVN Excel Diff Tool")
    parser.add_argument("directory", help="SVN working copy directory path")
    parser.add_argument("--file", help="Diff a specific file only", default=None)
    parser.add_argument("--json", action="store_true", help="Output as JSON")
    parser.add_argument("--html", action="store_true", help="Generate HTML diff and open in browser")
    args = parser.parse_args()

    directory = os.path.expanduser(args.directory)

    if not os.path.exists(directory):
        print(f"Error: path does not exist: {directory}", file=sys.stderr)
        sys.exit(1)

    if args.file:
        filepath = args.file if os.path.isabs(args.file) else os.path.join(directory, args.file)
        changes = [("M", filepath)]
    else:
        changes = get_svn_status(directory)

    if not changes:
        print("No changes detected.")
        sys.exit(0)

    excel_exts = {".xls", ".xlsx"}
    results = []  # list of (filepath, status, unified_diff, text_diff)

    for status, filepath in changes:
        ext = os.path.splitext(filepath)[1].lower()

        if status == "?":
            continue

        if status in ("A", "D"):
            results.append((filepath, status, None, None))
            continue

        if ext in excel_exts:
            base_tmp = get_svn_base(filepath)
            if base_tmp is None:
                continue
            try:
                base_data = read_excel_to_rows(base_tmp)
                work_data = read_excel_to_rows(filepath)
                udiff = build_unified_diff(base_data, work_data)
                results.append((filepath, status, udiff, None))
            finally:
                os.unlink(base_tmp)
        else:
            r = _run_subprocess([get_svn_command(), "diff", filepath], capture_output=True, text=True)
            if r.stdout.strip():
                results.append((filepath, status, None, r.stdout))

    if not results:
        print("No changes detected.")
        sys.exit(0)

    if args.html:
        html_content = generate_html(results, directory)
        out_path = os.path.join(
            tempfile.gettempdir(),
            f"svn_diff_{os.path.basename(directory)}_{datetime.now().strftime('%H%M%S')}.html",
        )
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        webbrowser.open(f"file://{out_path}")
        print(f"Diff report opened in browser: {out_path}")
    elif args.json:
        json_data = {}
        for filepath, status, udiff, text_diff in results:
            if udiff:
                # Convert sets to lists for JSON
                serializable = {}
                for sn, sd in udiff.items():
                    rows = []
                    for r in sd["rows"]:
                        rc = dict(r)
                        rc["changed_cols"] = list(rc["changed_cols"])
                        rows.append(rc)
                    serializable[sn] = {"headers": sd["headers"], "rows": rows}
                json_data[filepath] = serializable
        print(json.dumps(json_data, ensure_ascii=False, indent=2, default=str))
    else:
        reports = []
        for filepath, status, udiff, text_diff in results:
            if status in ("A", "D"):
                reports.append(format_report(filepath, None, is_new=(status == "A"), is_deleted=(status == "D")))
            elif udiff:
                report = format_report(filepath, udiff)
                if report:
                    reports.append(report)
            elif text_diff:
                rel_name = os.path.basename(filepath)
                reports.append(f"### [M] {rel_name}\n\n```diff\n{text_diff}\n```\n")

        if reports:
            print(f"## SVN Diff Report: `{os.path.basename(directory)}`\n")
            print(f"Changed files: {len(reports)}\n---\n")
            print("\n---\n".join(reports))


if __name__ == "__main__":
    main()
